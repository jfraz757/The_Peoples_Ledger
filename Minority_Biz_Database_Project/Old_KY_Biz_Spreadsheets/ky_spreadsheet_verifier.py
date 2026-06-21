"""
Kentucky Minority Business Spreadsheet Verifier
=================================================
Reads the 2023 MBDKY spreadsheet data, enriches minority type using
DBE ethnicity codes, verifies each business website is still active,
and outputs a cleaned CSV in the same format as ky_minority_businesses.csv.

Requirements:
    pip install requests beautifulsoup4 pandas openpyxl anthropic python-dotenv

Usage:
    1. Place your .xlsx files in the same folder as this script.
    2. Run: python ky_spreadsheet_verifier.py
    3. Output appends to ky_minority_businesses.csv

Input file expected:
    Darleen_s_February_Database_-_KCF_DEI_Website_2023.xlsx

Notes:
    - Businesses with out-of-state addresses are included but flagged.
    - Website check is a live HTTP request -- businesses with no website
      or a dead website are marked "Unverified" in the Status column.
    - Minority type is drawn first from the DBE ethnicity code, then
      from the Certification Type column, then from Claude if ambiguous.
"""

import os
import re
import time
import random
import hashlib
import requests
import pandas as pd
import openpyxl
from bs4 import BeautifulSoup
from anthropic import Anthropic
from dotenv import load_dotenv

# ─────────────────────────────────────────────
#  LOAD KEYS
# ─────────────────────────────────────────────
load_dotenv()
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")

# ─────────────────────────────────────────────
#  CONFIGURATION
# ─────────────────────────────────────────────
SCRIPT_DIR      = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE      = os.path.join(SCRIPT_DIR, "darleen_2023.xlsx")
OUTPUT_FILE     = r"C:\Users\jfraz\Desktop\Claude_KY_Biz_Databse\ky_minority_businesses.csv"
CHECKPOINT_FILE = os.path.join(SCRIPT_DIR, "checkpoint_verifier.csv")
CHECKPOINT_EVERY = 10
REQUEST_TIMEOUT  = 10
REQUEST_DELAY    = (1, 3)

# ─────────────────────────────────────────────
#  DBE ETHNICITY CODE --> MINORITY TYPE MAP
#  From the DBE sheet legend in your spreadsheet:
#  WF = White Female = WBE
#  BF = Black Female = Black-Owned, Women-Owned
#  BM = Black Male   = Black-Owned
#  HF = Hispanic Female = Latino-Owned, Women-Owned
#  HM = Hispanic Male   = Latino-Owned
#  NAF = Native American Female = Native American-Owned, Women-Owned
#  NAM = Native American Male   = Native American-Owned
#  SAF = Subcontinent Asian Female = Asian-Owned, Women-Owned
#  SAM = Subcontinent Asian Male   = Asian-Owned
#  APF = Asian Pacific Female = Asian-Owned, Women-Owned
#  APM = Asian Pacific Male   = Asian-Owned
#  AIF = American Indian Female = Native American-Owned, Women-Owned
#  AIM = American Indian Male   = Native American-Owned
# ─────────────────────────────────────────────
ETHNICITY_MAP = {
    "WF":  "Women-Owned",
    "BF":  "Black-Owned, Women-Owned",
    "BM":  "Black-Owned",
    "HF":  "Latino-Owned, Women-Owned",
    "HM":  "Latino-Owned",
    "NAF": "Native American-Owned, Women-Owned",
    "NAM": "Native American-Owned",
    "SAF": "Asian-Owned, Women-Owned",
    "SAM": "Asian-Owned",
    "APF": "Asian-Owned, Women-Owned",
    "APM": "Asian-Owned",
    "AIF": "Native American-Owned, Women-Owned",
    "AIM": "Native American-Owned",
}

# Certification type --> minority type fallback
CERT_TYPE_MAP = {
    "MBE":    "Minority-Owned (general)",
    "WBE":    "Women-Owned",
    "MWBE":   "Minority-Owned (general), Women-Owned",
    "DIBE":   "Disability-Owned",
    "LGBTBE": "LGBTQ+-Owned",
    "SDVOSB": "Veteran-Owned",
    "VOSB":   "Veteran-Owned",
}

# ─────────────────────────────────────────────
#  SETUP
# ─────────────────────────────────────────────
anthropic_client = Anthropic(api_key=ANTHROPIC_API_KEY)


def business_key(name: str, website: str) -> str:
    combined = f"{name.lower().strip()}|{website.lower().strip()}"
    return hashlib.md5(combined.encode()).hexdigest()


def polite_pause():
    time.sleep(random.uniform(*REQUEST_DELAY))


def normalize_url(url: str) -> str:
    """Ensures URL has a scheme so requests can fetch it."""
    if not url:
        return ""
    url = url.strip()
    if not url.startswith("http"):
        url = "https://" + url
    return url


# ─────────────────────────────────────────────
#  STEP 1: Read spreadsheet
# ─────────────────────────────────────────────
def load_spreadsheet() -> tuple[list[dict], dict[str, str]]:
    """
    Loads the master business list from 'Merging the Data' sheet
    and builds an ethnicity lookup from 'DBE - Delimited Firm Info'.
    Returns (records list, ethnicity_lookup dict).
    """
    print(f"Loading spreadsheet: {os.path.basename(INPUT_FILE)}")
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)

    # ── Build ethnicity lookup from DBE sheet ─────────────────────────
    ethnicity_lookup = {}
    try:
        ws_dbe = wb["DBE - Delimited Firm Info"]
        rows_dbe = list(ws_dbe.iter_rows(values_only=True))
        for row in rows_dbe[1:]:
            firm_name = str(row[0]).strip().upper() if row[0] else ""
            ethnicity  = str(row[1]).strip().upper() if row[1] else ""
            if firm_name and ethnicity:
                ethnicity_lookup[firm_name] = ethnicity
        print(f"  Loaded {len(ethnicity_lookup)} ethnicity codes from DBE sheet")
    except Exception as e:
        print(f"  Could not load DBE ethnicity sheet: {e}")

    # ── Load master business list ─────────────────────────────────────
    ws = wb["Merging the Data"]
    rows = list(ws.iter_rows(values_only=True))

    # Column mapping based on what we read earlier
    COL = {
        "name":      0,
        "cert_exp":  1,
        "cert_type": 2,
        "industry":  3,
        "services":  4,
        "owner":     5,
        "certifier": 6,
        "county":    7,
        "address":   8,
        "city":      9,
        "state":     10,
        "zip":       11,
        "phone":     12,
        "email":     13,
        "website":   14,
        "naics":     15,
    }

    records = []
    for row in rows[1:]:
        name = str(row[COL["name"]]).strip() if row[COL["name"]] else ""
        if not name or name.lower() == "none":
            continue

        # Extract certification type -- clean the long header text
        cert_raw = str(row[COL["cert_type"]]).strip() if row[COL["cert_type"]] else ""
        # Pull just the code (MBE, WBE, etc.) from the cell
        cert_match = re.search(
            r"\b(MBE|WBE|MWBE|DIBE|LGBTBE|SDVOSB|VOSB)\b",
            cert_raw, re.IGNORECASE
        )
        cert_type = cert_match.group(1).upper() if cert_match else cert_raw[:20]

        # Determine minority type from ethnicity code first, then cert type
        name_upper = name.upper()
        ethnicity_code = ethnicity_lookup.get(name_upper, "")
        if ethnicity_code and ethnicity_code in ETHNICITY_MAP:
            minority_type = ETHNICITY_MAP[ethnicity_code]
        elif cert_type in CERT_TYPE_MAP:
            minority_type = CERT_TYPE_MAP[cert_type]
        else:
            minority_type = ""

        # Build address
        street  = str(row[COL["address"]]).strip() if row[COL["address"]] else ""
        city    = str(row[COL["city"]]).strip()    if row[COL["city"]]    else ""
        state   = str(row[COL["state"]]).strip()   if row[COL["state"]]   else ""
        zip_    = str(row[COL["zip"]]).strip()     if row[COL["zip"]]     else ""
        address = ", ".join(filter(None, [street, city, state, zip_]))

        # Clean services -- strip NAICS codes from front, cap length
        services_raw = str(row[COL["services"]]).strip() if row[COL["services"]] else ""
        services = services_raw[:300] if services_raw else ""

        # Flag out-of-state businesses
        is_ky = state.upper() in ("KY", "KENTUCKY") or not state

        records.append({
            "business_name": name,
            "address":       address,
            "phone":         str(row[COL["phone"]]).strip() if row[COL["phone"]] else "",
            "services":      services,
            "website":       normalize_url(str(row[COL["website"]]).strip() if row[COL["website"]] else ""),
            "minority_type": minority_type,
            "cert_type":     cert_type,
            "certifier":     str(row[COL["certifier"]]).strip() if row[COL["certifier"]] else "",
            "owner":         str(row[COL["owner"]]).strip() if row[COL["owner"]] else "",
            "is_ky":         is_ky,
            "status":        "Pending",
        })

    wb.close()
    print(f"  Loaded {len(records)} businesses from master sheet")
    ky_count = sum(1 for r in records if r["is_ky"])
    print(f"  {ky_count} Kentucky-based, {len(records) - ky_count} out-of-state")
    return records, ethnicity_lookup


# ─────────────────────────────────────────────
#  STEP 2: Verify website is still live
# ─────────────────────────────────────────────
def check_website(url: str) -> str:
    """
    Returns 'Active', 'Inactive', or 'No Website'.
    Active = HTTP 200 or 301/302 redirect that resolves.
    Inactive = connection error, timeout, 404, 410, or parked domain.
    """
    if not url:
        return "No Website"
    try:
        resp = requests.get(
            url,
            timeout=REQUEST_TIMEOUT,
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"},
            allow_redirects=True,
        )
        # Check for parked domain / expired page signals
        if resp.status_code in (404, 410, 500, 503):
            return "Inactive"

        # Check page content for domain-parking signals
        text = resp.text.lower()
        parked_signals = [
            "this domain is for sale", "domain is parked",
            "buy this domain", "website coming soon",
            "under construction", "account suspended",
            "this site can't be reached",
        ]
        if any(sig in text for sig in parked_signals):
            return "Inactive"

        return "Active"

    except requests.exceptions.SSLError:
        # Try http fallback
        try:
            url_http = url.replace("https://", "http://")
            resp = requests.get(url_http, timeout=REQUEST_TIMEOUT,
                                headers={"User-Agent": "Mozilla/5.0"})
            return "Active" if resp.status_code == 200 else "Inactive"
        except Exception:
            return "Inactive"
    except Exception:
        return "Inactive"


# ─────────────────────────────────────────────
#  STEP 3: Claude fills minority type if missing
# ─────────────────────────────────────────────
FILL_PROMPT = """
You are helping build a minority-owned business database.
A business record is missing its minority ownership type.
Based on the information below, determine the most likely minority type.

Business Name: {name}
Owner: {owner}
Certification Type: {cert_type}
Certifier: {certifier}
Services: {services}

Return ONLY one of these labels (comma-separate if multiple apply):
Black-Owned, Latino-Owned, Asian-Owned, Native American-Owned,
Women-Owned, LGBTQ+-Owned, Veteran-Owned, Disability-Owned,
Muslim-Owned, Minority-Owned (general)

If you cannot determine the minority type from the information given,
return exactly: Unknown

Return only the label(s), nothing else.
"""

def fill_minority_type_with_claude(record: dict) -> str:
    """Uses Claude to infer minority type when not determinable from codes."""
    try:
        prompt = FILL_PROMPT.format(
            name      = record["business_name"],
            owner     = record["owner"],
            cert_type = record["cert_type"],
            certifier = record["certifier"],
            services  = record["services"][:200],
        )
        response = anthropic_client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=100,
            messages=[{"role": "user", "content": prompt}],
        )
        result = response.content[0].text.strip()
        return result if result != "Unknown" else "Minority-Owned (general)"
    except Exception as e:
        print(f"  [Claude Error] {e}")
        return "Minority-Owned (general)"


# ─────────────────────────────────────────────
#  STEP 4: Load checkpoint
# ─────────────────────────────────────────────
def load_checkpoint() -> tuple[list, set]:
    if not os.path.exists(CHECKPOINT_FILE):
        return [], set()
    try:
        df = pd.read_csv(CHECKPOINT_FILE, encoding="utf-8-sig")
        records = df.rename(columns={
            "Business Name":       "business_name",
            "Address":             "address",
            "Phone":               "phone",
            "Services / Products": "services",
            "Website":             "website",
            "Minority Type":       "minority_type",
            "Status":              "status",
            "Kentucky Based":      "is_ky",
        }).to_dict("records")
        seen = {business_key(r["business_name"], r.get("website", "")) for r in records}
        print(f"  [Resumed from checkpoint: {len(records)} businesses already processed]")
        return records, seen
    except Exception as e:
        print(f"  [Could not load checkpoint: {e}]")
        return [], set()


# ─────────────────────────────────────────────
#  STEP 5: Save CSV
# ─────────────────────────────────────────────
def save_csv(data: list[dict], filename: str):
    """Saves in the same format as ky_minority_businesses.csv."""
    rows = []
    for d in data:
        rows.append({
            "Business Name":       d.get("business_name", ""),
            "Address":             d.get("address", ""),
            "Phone":               d.get("phone", ""),
            "Services / Products": d.get("services", ""),
            "Website":             d.get("website", ""),
            "Minority Type":       d.get("minority_type", ""),
            "Status":              d.get("status", ""),
            "Kentucky Based":      "Yes" if d.get("is_ky") else "No",
        })
    df = pd.DataFrame(rows)
    df.drop_duplicates(subset=["Business Name", "Website"], inplace=True)
    df.sort_values("Business Name", inplace=True)
    df.to_csv(filename, index=False, encoding="utf-8-sig")


def merge_into_main_csv(verified_data: list[dict]):
    """
    Merges verified records into the main ky_minority_businesses.csv.

    Duplicate handling:
    - Matches are found by business name + website key.
    - When a duplicate is found, the spreadsheet record WINS on minority_type
      since it comes from certified government data, which is more authoritative
      than what the web scraper inferred from page text.
    - All other fields (address, phone, services) are updated from the
      spreadsheet if the existing record has them blank.
    - Net-new businesses from the spreadsheet are added to the CSV.
    """
    # Build a lookup of existing scraper records by business key
    existing_by_key = {}

    if os.path.exists(OUTPUT_FILE):
        try:
            df_existing = pd.read_csv(OUTPUT_FILE, encoding="utf-8-sig")
            for _, row in df_existing.iterrows():
                name = str(row.get("Business Name", ""))
                site = str(row.get("Website", ""))
                bkey = business_key(name, site)
                existing_by_key[bkey] = row.to_dict()
            print(f"  Existing main CSV has {len(existing_by_key)} records")
        except Exception as e:
            print(f"  Could not read existing CSV: {e}")

    added    = 0
    updated  = 0

    for d in verified_data:
        bkey = business_key(d.get("business_name", ""), d.get("website", ""))

        if bkey in existing_by_key:
            # Duplicate found -- spreadsheet minority type wins
            existing = existing_by_key[bkey]
            old_type = existing.get("Minority Type", "")
            new_type = d.get("minority_type", "")

            if new_type and new_type != old_type:
                existing["Minority Type"] = new_type
                print(f"  → Updated minority type: {d['business_name']}")
                print(f"    Scraper had: '{old_type}' → Spreadsheet says: '{new_type}'")
                updated += 1

            # Fill any blank fields from spreadsheet data
            if not existing.get("Address") and d.get("address"):
                existing["Address"] = d["address"]
            if not existing.get("Phone") and d.get("phone"):
                existing["Phone"] = d["phone"]
            if not existing.get("Services / Products") and d.get("services"):
                existing["Services / Products"] = d["services"]

            # Add Status and Kentucky Based if not present
            existing.setdefault("Status", d.get("status", ""))
            existing.setdefault("Kentucky Based", "Yes" if d.get("is_ky") else "No")

            existing_by_key[bkey] = existing

        else:
            # Net-new business from spreadsheet -- add it
            existing_by_key[bkey] = {
                "Business Name":       d.get("business_name", ""),
                "Address":             d.get("address", ""),
                "Phone":               d.get("phone", ""),
                "Services / Products": d.get("services", ""),
                "Website":             d.get("website", ""),
                "Minority Type":       d.get("minority_type", ""),
                "Status":              d.get("status", ""),
                "Kentucky Based":      "Yes" if d.get("is_ky") else "No",
            }
            added += 1

    df_merged = pd.DataFrame(list(existing_by_key.values()))
    df_merged.drop_duplicates(subset=["Business Name", "Website"], inplace=True)
    df_merged.sort_values("Business Name", inplace=True)
    df_merged.to_csv(OUTPUT_FILE, index=False, encoding="utf-8-sig")
    print(f"  Added {added} new records to {OUTPUT_FILE}")
    print(f"  Updated minority type on {updated} existing records")
    print(f"  Total records in main CSV: {len(df_merged)}")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
def verify_database():
    if not os.path.exists(INPUT_FILE):
        print(f"ERROR: Input file not found: {INPUT_FILE}")
        print("Make sure Darleen_s_February_Database_-_KCF_DEI_Website_2023.xlsx")
        print("is in the same folder as this script.")
        return

    # Load spreadsheet
    records, _ = load_spreadsheet()

    # Resume from checkpoint
    database, seen_businesses = load_checkpoint()

    print(f"\n=== Verifying {len(records)} businesses ===\n")

    active_count   = 0
    inactive_count = 0
    no_site_count  = 0

    for i, record in enumerate(records, 1):
        bkey = business_key(record["business_name"], record["website"])
        if bkey in seen_businesses:
            continue

        print(f"[{i}/{len(records)}] {record['business_name']}")

        # Step A: Fill missing minority type with Claude
        if not record["minority_type"]:
            print(f"  → No minority type found, asking Claude...")
            record["minority_type"] = fill_minority_type_with_claude(record)
            print(f"  → Claude says: {record['minority_type']}")

        # Step B: Check website status
        status = check_website(record["website"])
        record["status"] = status

        if status == "Active":
            active_count += 1
            print(f"  → Website: Active | Type: {record['minority_type']}")
        elif status == "Inactive":
            inactive_count += 1
            print(f"  → Website: Inactive | Type: {record['minority_type']}")
        else:
            no_site_count += 1
            print(f"  → No website on file | Type: {record['minority_type']}")

        seen_businesses.add(bkey)
        database.append(record)

        # Checkpoint save
        if len(database) % CHECKPOINT_EVERY == 0:
            save_csv(database, CHECKPOINT_FILE)
            print(f"  [Checkpoint: {len(database)} processed]")

        polite_pause()

    # ── Final save ────────────────────────────────────────────────────
    print(f"\n=== Verification Complete ===")
    print(f"  Active websites:   {active_count}")
    print(f"  Inactive websites: {inactive_count}")
    print(f"  No website on file: {no_site_count}")

    # Save standalone verifier output
    verifier_output = os.path.join(SCRIPT_DIR, "ky_verified_2023_businesses.csv")
    save_csv(database, verifier_output)
    print(f"\n  Standalone output: {verifier_output}")

    # Merge into main database
    print(f"\n=== Merging into main database ===")
    merge_into_main_csv(database)

    # Clean up checkpoint
    if os.path.exists(CHECKPOINT_FILE):
        os.remove(CHECKPOINT_FILE)
        print("  Checkpoint file removed")


# ─────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────
if __name__ == "__main__":
    if not ANTHROPIC_API_KEY:
        print("ERROR: ANTHROPIC_API_KEY not found in .env file")
        print("Add your key to .env and try again.")
    else:
        verify_database()
