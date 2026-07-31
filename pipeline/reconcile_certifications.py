"""
Lane 2: reconcile the Kentucky certifier lists against the live `businesses` table.

    python pipeline/reconcile_certifications.py                 # DRY RUN (default)
    python pipeline/reconcile_certifications.py --apply         # backfill labels only
    python pipeline/reconcile_certifications.py --apply --insert-new   # also add new businesses
    python pipeline/reconcile_certifications.py --inspect FILE  # describe one export

Reads the files staged by Minority_Biz_Database_Project/data_gather_.ipynb (which handles
the download/convert/rename step -- do not rebuild that here).

TWO JOBS
  1. BACKFILL. A business already in the directory that appears on a certifier list gets
     its certification_type filled. `certification_type` may ONLY come from one of the
     three certifying bodies -- never scraped, never inferred, never from a submitter.
     Blank remains correct for the large majority of the directory.
  2. INSERT. A certified business the directory does not have is added. These are high
     quality: state-verified, and often invisible to Lane 1 because they have no
     consumer web presence.

THE RISK THIS SCRIPT IS BUILT AROUND
The June 2026 merge of HRC data was done by hand and produced ~76 duplicate clusters;
dedupe_live.py had to remove 77 rows. The danger here is NOT a wrong certification value,
it is duplicate insertion. So:
  * matching is deliberately conservative, and every tier is reported before anything writes
  * insertion is behind a SECOND flag (--insert-new), separate from --apply
  * anything ambiguous goes to a review CSV instead of being guessed at

LABELS (verified against the 2026-07-29 exports)
  Louisville HRC   MBE, WBE, SDVOSB, VOSB, DIBE, LGBTBE
  KY Transportation DBE, SBE
  KY Finance       MBE, WBE, MWBE

SBE is size-based, not ownership-based. It is still recorded -- it is a real government
certification and 87 businesses already in the directory hold it -- but a business is
never ADDED on ownership grounds because of it.

CERTIFICATIONS NEVER ENTER minority_type
`certification_type` and `minority_type` are separate fields holding separate facts, and
this script must not blur them. The first version wrote "Minority-Owned" into minority_type
for an MBE holder with no recorded group, and "Small Business Enterprise" for an SBE-only
business. Both were wrong: the certification already states that fact in its own column,
and the RPC matches ownership with ilike '%value%', so a "Minority-Owned" ownership pill
returned 360 rows -- 77 government-verified MBE holders mixed with 283 unrelated
"Minority-Owned (general)" rows from the scraper's organic lane. Two vocabularies in one
column cannot be filtered apart. A blank minority_type is the correct answer when the
demographic is unknown; the business is still reachable through the state certification
filter, which reads the column where that fact actually lives.
"""

import argparse
import csv
import io
import json
import os
import re
import sys
import urllib.request
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from rapidfuzz import fuzz, process

PIPELINE_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(PIPELINE_DIR)
DATA_DIR = os.path.join(REPO_ROOT, "data")
SPREADSHEET_DIR = os.path.join(REPO_ROOT, "Minority_Biz_Database_Project", "Spreadsheets")

try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(REPO_ROOT, ".env"))
except ImportError:
    pass

SUPABASE_URL = os.getenv("SUPABASE_URL", "")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")
if not SUPABASE_KEY:
    raise SystemExit(
        "SUPABASE_SERVICE_ROLE_KEY missing from .env. This script writes to `businesses`, "
        "which anon can no longer do (restrict_anon_grants.sql, July 2026)."
    )
if "ursmecdpgtqckacyhnko" not in SUPABASE_URL:
    raise SystemExit(f"Refusing to run against an unexpected project: {SUPABASE_URL}")

# Certifications that reflect WHO OWNS the business. SBE is intentionally absent.
OWNERSHIP_CERTS = {"MBE", "WBE", "MWBE", "DBE", "VOSB", "SDVOSB", "DIBE", "LGBTBE"}
SIZE_CERTS = {"SBE"}
VALID_CERTS = OWNERSHIP_CERTS | SIZE_CERTS

# HRC's Ethnicity -> the site's ownership pills. "Caucasian" maps to nothing: those rows
# are women-owned businesses whose certification carries no racial-minority component,
# so the WBE label is the whole story.
ETHNICITY_TO_TYPE = {
    "african american": "Black-Owned",
    "hispanic american": "Latine-Owned",
    "asian-pacific american": "Asian-Owned",
    "subcontinent asian american": "Asian-Owned",
    "asian": "Asian-Owned",
    "native american": "Native American-Owned",
    "caucasian": None,
}
# Ownership implied by the certification itself, when the source has no Ethnicity column.
CERT_TO_TYPE = {
    "WBE": "Women-Owned",
    "MWBE": "Women-Owned",
    "VOSB": "Veteran-Owned",
    "SDVOSB": "Veteran-Owned, Disability-Owned",
    "DIBE": "Disability-Owned",
    "LGBTBE": "LGBTQ+-Owned",
}
# (removed July 2026: see derive_minority_type -- certifications never enter minority_type)

SUFFIX = re.compile(r"\b(llc|inc|incorporated|corp|corporation|company|co|ltd|llp|pllc|plc|lp)\b\.?")


def match_key(s):
    """Same normalisation prepare.py uses, so all three stages agree on identity."""
    s = (s or "").lower().replace("&", " and ")
    s = re.sub(r"[-/,]", " ", s)
    s = re.sub(r"[^a-z0-9 ]", "", s)
    s = SUFFIX.sub(" ", s)
    return re.sub(r"\s+", " ", s).strip()


def norm_site(u):
    return re.sub(r"^https?://(www\.)?", "", str(u or "").lower().strip()).rstrip("/")


def clean(v):
    if v is None:
        return ""
    s = str(v).replace("\t", " ").replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return "" if s.lower() in ("none", "n/a", "na", "null", "nan") else s


def read_b2gnow(path):
    """Both B2GNow exports (HRC, KYTC) share one layout: a preamble, then the header.

    cp1252, not UTF-8 -- they contain 0x96 en-dashes and a UTF-8 read raises outright.
    The header row is located rather than assumed, because the preamble length varies.
    """
    raw = Path(path).read_bytes()
    text = None
    for enc in ("utf-8-sig", "cp1252"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = raw.decode("cp1252", errors="replace")
    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    hdr = next((i for i, l in enumerate(lines[:40]) if l.lstrip('"').startswith("Company Name")), None)
    if hdr is None:
        raise SystemExit(f"{Path(path).name}: no 'Company Name' header in the first 40 lines.")
    # DictReader keeps only the LAST of a repeated column, and B2GNow emits City/State/Zip
    # twice (physical then mailing) -- which would silently make the MAILING address win.
    raw_header = next(csv.reader(io.StringIO(lines[hdr])))
    seen, names = defaultdict(int), []
    for c in raw_header:
        c = c.strip()
        if not c:
            continue
        seen[c] += 1
        names.append(c if seen[c] == 1 else f"{c} (mailing)")
    out = []
    for vals in csv.reader(io.StringIO("\n".join(lines[hdr + 1:]))):
        if not vals or not clean(vals[0]):
            continue
        out.append({names[i]: clean(v) for i, v in enumerate(vals) if i < len(names)})
    return out


def read_mwbe_xlsx(path):
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("openpyxl required for .xlsx. pip install openpyxl")
    import warnings
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    grid = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    hdr = next((i for i, r in enumerate(grid[:12]) if r and clean(r[0]) == "Name"), None)
    if hdr is None:
        raise SystemExit(f"{Path(path).name}: no 'Name' header row.")
    cols = [clean(c) for c in grid[hdr] if clean(c)]
    out = []
    for r in grid[hdr + 1:]:
        cells = [clean(c) for c in (r or ())]
        if not cells or not cells[0]:
            continue
        out.append({cols[i]: cells[i] for i in range(min(len(cols), len(cells)))})
    return out


def find_source_files():
    """Locate the staged exports. Newest file wins if a folder holds several."""
    found = {}
    for label, folder, ext in (
        ("Louisville HRC", "Louisville_HRC", ".csv"),
        ("KY Transportation", "KY_Transportation_Cabinet", ".csv"),
        ("KY Finance", "KY_Finance_&_Administration", (".xlsx", ".csv")),
    ):
        d = Path(SPREADSHEET_DIR) / folder
        if not d.exists():
            continue
        exts = (ext,) if isinstance(ext, str) else ext
        files = [p for p in d.iterdir() if p.is_file() and p.suffix.lower() in exts]
        if files:
            found[label] = max(files, key=lambda p: p.stat().st_mtime)
    return found


def load_certifiers():
    """-> {match_key: {name, labels{label:source}, ethnicity, address, city, phone, website}}"""
    files = find_source_files()
    if not files:
        raise SystemExit(f"No certifier files found under {SPREADSHEET_DIR}")

    recs = {}

    def add(name, label, source, **fields):
        label = clean(label).upper()
        if label not in VALID_CERTS:
            return
        k = match_key(name)
        if not k:
            return
        r = recs.setdefault(k, {"name": clean(name), "labels": {}, "ethnicity": "",
                                "address": "", "city": "", "phone": "", "website": ""})
        r["labels"][label] = source
        # First non-empty value wins; HRC is loaded first and is the richest source.
        for f, v in fields.items():
            if v and not r.get(f):
                r[f] = v

    for label, path in files.items():
        if label == "Louisville HRC":
            for row in read_b2gnow(path):
                if row.get("State", "").upper() != "KY":
                    continue
                add(row["Company Name"], row.get("Certification Type"), "HRC",
                    ethnicity=row.get("Ethnicity", ""), address=row.get("Physical Address", ""),
                    city=row.get("City", ""), phone=row.get("Phone", ""),
                    website=row.get("Website", ""))
        elif label == "KY Transportation":
            for row in read_b2gnow(path):
                if row.get("State", "").upper() != "KY":
                    continue
                add(row["Company Name"], row.get("Certification Type"), "KYTC",
                    address=row.get("Physical Address", ""), city=row.get("City", ""),
                    phone=row.get("Phone", ""), website=row.get("Website", ""))
        elif label == "KY Finance":
            rows = read_mwbe_xlsx(path) if path.suffix.lower() == ".xlsx" else read_b2gnow(path)
            for row in rows:
                if str(row.get("State", "")).upper() != "KY":
                    continue
                add(row.get("Name") or row.get("Company Name"), row.get("Type (*)"), "KYFIN",
                    city=row.get("City", ""), phone=row.get("Phone #", ""),
                    website=row.get("Web", ""))
    return recs, files


def fetch_live():
    rows, off = [], 0
    while True:
        req = urllib.request.Request(
            f"{SUPABASE_URL}/rest/v1/businesses?select=id,business_name,website,minority_type,"
            f"certification_type,address&order=id.asc&limit=1000&offset={off}",
            headers={"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"})
        page = json.loads(urllib.request.urlopen(req, timeout=60).read())
        rows += page
        if len(page) < 1000:
            break
        off += 1000
    return rows


def derive_minority_type(rec):
    """Ownership DEMOGRAPHIC implied by the certification data, or None.

    Returns only values that exist as ownership pills on the site. It deliberately does
    NOT invent a placeholder when the demographic is unknown.

    The first version of this returned "Minority-Owned" for an MBE holder whose certifier
    did not record a group, and "Small Business Enterprise" for an SBE-only business. Both
    were wrong for the same reason: they put CERTIFICATION vocabulary into the ownership
    column. `certification_type` is a separate field and already carries that fact --
    writing it twice, in two vocabularies, in one column is what created the collision.
    The RPC matches ownership with ilike '%value%', so a "Minority-Owned" pill returned
    360 rows: 77 government-verified MBE holders plus the 283 unrelated
    "Minority-Owned (general)" rows from the scraper's organic lane. Nothing could tell
    them apart, because they were never the same kind of claim.

    A blank ownership is the honest answer when the demographic is genuinely unknown.
    Those businesses are still findable -- by name, by industry, and by the state
    certification filter, which reads certification_type where the fact actually lives.
    """
    types = []
    eth = ETHNICITY_TO_TYPE.get(clean(rec.get("ethnicity")).lower(), None)
    if eth:
        types.append(eth)
    for lab in rec["labels"]:
        mapped = CERT_TO_TYPE.get(lab)
        if mapped:
            types.extend(p.strip() for p in mapped.split(","))
    seen, out = set(), []
    for t in types:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return ", ".join(out) or None


def patch(bid, payload):
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/businesses?id=eq.{bid}",
        data=json.dumps(payload).encode(),
        headers={"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}",
                 "Content-Type": "application/json", "Prefer": "return=minimal"},
        method="PATCH")
    urllib.request.urlopen(req, timeout=30).read()


def insert(batch):
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/businesses",
        data=json.dumps(batch).encode(),
        headers={"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}",
                 "Content-Type": "application/json", "Prefer": "return=minimal"},
        method="POST")
    urllib.request.urlopen(req, timeout=60).read()


def main():
    ap = argparse.ArgumentParser(description="Lane 2: reconcile certifier lists against the live table.")
    ap.add_argument("--apply", action="store_true", help="write certification_type backfills")
    ap.add_argument("--insert-new", action="store_true",
                    help="also insert certified businesses the directory does not have "
                         "(requires --apply)")
    ap.add_argument("--fuzzy", type=int, default=93,
                    help="fuzzy match cutoff for BACKFILL (default 93). Matches below this "
                         "go to the review CSV, never applied automatically.")
    args = ap.parse_args()

    recs, files = load_certifiers()
    print("Source files:")
    for lab, p in files.items():
        print(f"  {lab:<18} {p.name}")
    print(f"\nKY certified companies across all sources: {len(recs)}")

    live = fetch_live()
    by_key = {}
    for b in live:
        k = match_key(b["business_name"])
        if k:
            by_key.setdefault(k, b)
    sites = {norm_site(b["website"]): b for b in live if b.get("website")}
    sites.pop("", None)
    choices = list(by_key.keys())
    print(f"Live directory: {len(live)} rows, {len(by_key)} distinct name keys\n")

    backfill, inserts, review = [], [], []
    for k, rec in recs.items():
        hit = by_key.get(k) or sites.get(norm_site(rec.get("website")))
        method = "exact" if hit else ""
        if not hit:
            m = process.extractOne(k, choices, scorer=fuzz.token_sort_ratio, score_cutoff=args.fuzzy)
            if m:
                hit, method = by_key[m[0]], f"fuzzy {m[1]:.0f}"
            else:
                # A near-miss is the duplicate-insertion risk. Anything in this band is
                # neither backfilled nor inserted -- it goes to a human.
                near = process.extractOne(k, choices, scorer=fuzz.token_sort_ratio, score_cutoff=85)
                if near:
                    review.append((rec, by_key[near[0]], near[1]))
                    continue
        if hit:
            have = {c.strip().upper() for c in str(hit.get("certification_type") or "").split(",") if c.strip()}
            add = set(rec["labels"]) - have
            if add:
                backfill.append((hit, rec, sorted(have | set(rec["labels"])), method))
        else:
            inserts.append(rec)

    own = [r for r in inserts if set(r["labels"]) & OWNERSHIP_CERTS]
    sbe = [r for r in inserts if not (set(r["labels"]) & OWNERSHIP_CERTS)]
    print(f"BACKFILL certification_type on existing rows : {len(backfill)}")
    print(f"INSERT   certified businesses not in directory: {len(inserts)}"
          f"   ({len(own)} ownership-certified, {len(sbe)} SBE-only)")
    print(f"REVIEW   ambiguous name matches (85-{args.fuzzy}) : {len(review)}   <- never auto-applied")

    if review:
        out = os.path.join(DATA_DIR, f"cert_review_{datetime.now():%Y-%m-%d_%H%M%S}.csv")
        with open(out, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["certifier_name", "labels", "possible_live_match", "live_id", "score"])
            for rec, hit, sc in review:
                w.writerow([rec["name"], ",".join(sorted(rec["labels"])),
                            hit["business_name"], hit["id"], f"{sc:.0f}"])
        print(f"  -> {os.path.basename(out)}")

    print("\n  sample backfills:")
    for hit, rec, merged, method in backfill[:8]:
        print(f"    id={hit['id']:<5} {hit['business_name'][:34]:36} + {','.join(sorted(set(rec['labels'])))}  [{method}]")
    print("\n  sample inserts:")
    for rec in inserts[:8]:
        print(f"    {rec['name'][:38]:40} {','.join(sorted(rec['labels'])):<14} -> {derive_minority_type(rec)}")

    if not args.apply:
        print("\nDRY RUN. Nothing written. Re-run with --apply (and --insert-new to add businesses).")
        return

    n = 0
    for hit, rec, merged, method in backfill:
        patch(hit["id"], {"certification_type": ", ".join(merged)})
        n += 1
        if n % 50 == 0:
            print(f"  backfilled {n}/{len(backfill)}")
    print(f"\nBackfilled {n} row(s).")

    if args.insert_new:
        payload = []
        for rec in inserts:
            payload.append({
                "business_name": rec["name"],
                "address": ", ".join(p for p in [rec.get("address"), rec.get("city"), "KY"] if p) or None,
                "phone": rec.get("phone") or None,
                "website": rec.get("website") or None,
                "minority_type": derive_minority_type(rec),
                "certification_type": ", ".join(sorted(rec["labels"])),
                "kentucky_based": "Yes",
                "status": "Active",
                "source": "certifier reconcile " + datetime.now().strftime("%Y-%m-%d"),
            })
        for i in range(0, len(payload), 100):
            insert(payload[i:i + 100])
            print(f"  inserted {min(i+100, len(payload))}/{len(payload)}")
        print(f"\nInserted {len(payload)} business(es).")
        print("Next: python pipeline/enrich.py  ->  node generate-business-pages.js")
    else:
        print("(--insert-new not given, so no new businesses were added.)")


if __name__ == "__main__":
    main()
